from flask import Flask, request, render_template, jsonify, send_file, flash, redirect, url_for, session
import numpy as np
import math
import pandas as pd
import io
from datetime import datetime
from flask_sqlalchemy import SQLAlchemy
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from flask_migrate import Migrate
import os
from werkzeug.utils import secure_filename
import os
from flask import send_from_directory
from flask_login import UserMixin
from flask_login import LoginManager, current_user, login_user, login_required
import plotly
import plotly.graph_objs as go
import json
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import webbrowser
import os
from flask import Flask, request

app = Flask(__name__, static_folder=os.path.abspath("static"))
app.secret_key = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///site.db'  # ou outro banco de dados que estiver usando
db = SQLAlchemy(app)
migrate = Migrate(app, db)
# Certifique-se de definir o caminho correto para onde deseja salvar as imagens
UPLOAD_FOLDER = os.path.join(os.getcwd(), r'C:\Users\nycol\motor_analysis\static')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # Define a rota para login

# Função de callback para carregar o usuário a partir do ID
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@app.template_filter('getattr')
def getattr_filter(obj, attr):
    return getattr(obj, attr, None)

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), nullable=False, unique=True)
    email = db.Column(db.String(150), nullable=False, unique=True)
    password = db.Column(db.String(150), nullable=False)
    profile_picture = db.Column(db.String(100), nullable=True)  # Campo de imagem de perfil
    ICONE1 = db.Column(db.String(100), nullable=True)  # Campo de imagem de perfil
    ICONE2 = db.Column(db.String(100), nullable=True)  # Campo de imagem de perfil
    CAPA = db.Column(db.String(100), nullable=True)  # Campo de imagem de perfil
    is_admin = db.Column(db.Boolean, default=False)

class Motor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(300), nullable=False)
    P = db.Column(db.Float, nullable=False)
    V = db.Column(db.Float, nullable=False)
    f = db.Column(db.Float, nullable=False)
    Fp = db.Column(db.Float, nullable=False)
    rend = db.Column(db.Float, nullable=False)
    Fpv = db.Column(db.Float, nullable=False)
    Wn = db.Column(db.Float, nullable=False)
    Iv = db.Column(db.Float, nullable=False)
    Vf = db.Column(db.Float, nullable=False)
    Pav = db.Column(db.Float, nullable=False)
    Pe= db.Column(db.Float, nullable=False)
    Jm= db.Column(db.Float, nullable=False)
    tipo_partida = db.Column(db.String(50), nullable=False)
    fator = db.Column(db.Float, nullable=True)  # Adicionar o campo 'fator'
    rs = db.Column(db.Float, nullable=False)
    xs = db.Column(db.Float, nullable=False)
    rr = db.Column(db.Float, nullable=False)
    xr = db.Column(db.Float, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='motores')

def round_values(value, decimal_places=4):
    """Arredonda um valor único ou lista de valores para o número especificado de casas decimais."""
    if value is None:
        return 0.0  # Ou outro valor padrão que faça sentido para você
    if isinstance(value, list) or isinstance(value, np.ndarray):
        return [round(v, decimal_places) if v is not None else 0.0 for v in value]
    return round(value, decimal_places)

def safe_value(value):
    """Retorna um valor seguro, tratando NaN, infinitos e arredondando"""
    if value is None or value == float('inf') or value == float('-inf') or value != value:  # NaN check
        return 0.0
    return round(value, 4)

def convert_complex(value):
    if isinstance(value, complex):
        return value.imag  # Retorna apenas a parte imaginária
    return value

def prepare_graph_data(raw_data):
    """Prepara os dados do gráfico arredondando e tratando valores seguros"""
    return [safe_value(value) for value in raw_data]

def filter_values(series):
    return series[series > 0].values

def create_tables():
    with app.app_context():
        db.create_all()

def create_master_user():
    with app.app_context():
        username = 'admin'
        email = 'admin@example.com'  # Adicione um email padrão
        password = 'admin123'

        # Verifica se o administrador já existe
        existing_user = User.query.filter_by(username=username).first()
        if not existing_user:
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
            new_admin = User(username=username, email=email, password=hashed_password, is_admin=True)
            db.session.add(new_admin)
            db.session.commit()
            print(f"Usuário master '{username}' criado com sucesso.")
        else:
            print(f"Usuário master '{username}' já existe.")


def run_server():
    """Função para iniciar o servidor Flask manualmente"""
    webbrowser.open("http://127.0.0.1:5000")  # Abre o navegador

@app.route('/shutdown')
def shutdown():
    """Fecha o servidor Flask quando o usuário sai"""
    os._exit(0)  # Mata o processo do servidor
    return "O servidor está sendo encerrado..."

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.is_admin and current_user.id != user_id:
        flash('Você não tem permissão para acessar esta página.', 'danger')
        return redirect(url_for('home'))

    user = User.query.get_or_404(user_id)  # Busca o usuário no banco de dados ou retorna 404
    if request.method == 'POST':
        try:
            user.username = request.form.get('username', user.username)
            user.email = request.form.get('email', user.email)
            user.is_admin = request.form.get('is_admin') == 'on' if current_user.is_admin else user.is_admin

            # Atualizar senha, se fornecida
            new_password = request.form.get('new_password')
            if new_password:
                user.password = generate_password_hash(new_password)

            # Atualizar a imagem de perfil, se fornecida
            profile_picture = request.files.get('profile_picture')
            if profile_picture:
                filename = secure_filename(profile_picture.filename)
                profile_picture.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                user.profile_picture = filename

            db.session.commit()  # Salva as alterações no banco de dados
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('manage_users') if current_user.is_admin else url_for('home'))

        except Exception as e:
            db.session.rollback()
            flash(f'Ocorreu um erro ao atualizar o usuário: {e}', 'danger')

    return render_template('Edit_user.html', user=user)

@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    CAPA = ICONE1 = ICONE2 = None  # Variáveis padrão

    # Obtém o primeiro usuário do banco de dados (ou defina uma lógica melhor)
    user = User.query.first()
    if user:
        CAPA = user.CAPA
        ICONE1 = user.ICONE1
        ICONE2 = user.ICONE2

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.password, password):
            login_user(user)
            flash('Login realizado com sucesso!')
            return redirect(url_for('home'))
        else:
            flash('Usuário ou senha incorretos.')
            return redirect(url_for('login'))

    return render_template('login.html', CAPA=CAPA, ICONE1=ICONE1, ICONE2=ICONE2)

@app.route('/gerenciador_motor')
@login_required
def gerenciador_motor():
    user = current_user  # Pega o usuário logado diretamente
    user_profile_image_url = user.profile_picture
    username = user.username
    # Recupera o primeiro motor do usuário logado (ou um específico, conforme lógica)
    motor = Motor.query.filter_by(user_id=current_user.id).first()
    
    if not motor:
        flash('Nenhum motor encontrado. Por favor, crie um motor primeiro.', 'info')
        return redirect(url_for('motor'))
    return render_template('gerenciador_motor.html', motor=motor, user=user, user_profile_image_url=user_profile_image_url, username=username)

@app.route('/delete_motor', methods=['GET', 'POST'])
@login_required
def deletar_motor():
    try:
        # Recuperar o usuário logado
        user = current_user

        if request.method == 'POST':
            # Recuperar o ID do motor a partir do formulário
            motor_id = request.form.get('motor_id')
            if not motor_id:
                flash('Nenhum motor selecionado.', 'warning')
                return redirect(url_for('deletar_motor'))
            
            # Localizar o motor no banco de dados
            motor = Motor.query.filter_by(id=motor_id, user_id=user.id).first()
            if motor:
                db.session.delete(motor)  # Deletar motor
                db.session.commit()
                flash('Motor deletado com sucesso!', 'success')
            else:
                flash('Motor não encontrado ou não pertence a você.', 'danger')
            return redirect(url_for('deletar_motor'))

        # Recuperar todos os motores pertencentes ao usuário
        motores = Motor.query.filter_by(user_id=user.id).all()
        user_profile_image_url = user.profile_picture
        username = user.username

        return render_template('delete_motor.html', motors=motores, user=user, user_profile_image_url=user_profile_image_url, username=username)
    except Exception as e:
        flash(f"Erro: {e}", 'danger')
        return redirect(url_for('home'))

@app.route('/motor', methods=['GET', 'POST'])
@login_required
def motor():
        # Recupera os motores do usuário logado
    motores = Motor.query.filter_by(user_id=current_user.id).all()
    user = current_user
    user_profile_image_url = user.profile_picture
    username = user.username
    
    if request.method == 'POST':

        # Processa os dados do formulário
        name = request.form['name']
        P = request.form['P']
        V = request.form['V']
        f = request.form['f']
        Fp = request.form['Fp']
        rend = request.form['rend']
        Jm = request.form['Jm']
        Fpv = request.form['Fpv']
        Wn = request.form['Wn']
        Iv = request.form['Iv']
        Vf = request.form['Vf']
        Pav = request.form['Pav']
        Pe= request.form['Pe']
        tipo_partida = request.form['tipo_partida']
        rs = request.form['rs']
        xs = request.form['xs']
        rr = request.form['rr']
        xr = request.form['xr']
        fator = request.form.get('fator')  # Captura o fator, se for preenchido

        # Adiciona o novo motor ao banco de dados
        novo_motor = Motor(
            name=name,
            P=P,
            V=V,
            f=f,
            Fp=Fp,
            rend=rend,
            fator=fator if fator else None,  # Se o fator não for preenchido, define como None
            Jm=Jm,  # Inclua o campo Jm corretamente
            Fpv=Fpv,
            Wn=Wn,
            Iv=Iv,
            Vf=Vf,
            Pav=Pav,
            Pe=Pe,
            tipo_partida=tipo_partida,
            rs=rs,
            xs=xs,
            rr=rr,
            xr=xr,
            user_id=current_user.id  # Relaciona com o usuário logado
        )

        db.session.add(novo_motor)
        db.session.commit()

        flash('Motor cadastrado com sucesso!')
        return redirect(url_for('motor'))


    return render_template('Cadastro_motor.html',user=user, motor=motor, username=username, user_profile_image_url=user_profile_image_url)

@app.route('/home')
@login_required
def home():
    
    user = current_user
    user_profile_image_url = user.profile_picture
    username = user.username
    
    return render_template('home.html', user=user, user_profile_image_url=user_profile_image_url, username=username)

@app.route('/editar_motor', methods=['GET', 'POST'])
@login_required
def editar_motor():
    motores = Motor.query.filter_by(user_id=current_user.id).all()
    motor = None  # Variável para armazenar o motor selecionado
    user = current_user
    user_profile_image_url = user.profile_picture
    username = user.username
    # Se o método for POST, tenta editar o motor
    if request.method == 'POST':
        motor_id = request.form.get('motor_id')
        if motor_id:
            motor = Motor.query.get_or_404(motor_id)
        
        # Verifica se o motor existe e se pertence ao usuário
        if motor and motor.user_id == current_user.id:
            motor.name = request.form['name']
            motor.P = request.form['P']
            motor.V = request.form['V']
            motor.f = request.form['f']
            motor.Fp = request.form['Fp']
            motor.rend = request.form['rend']
            motor.Jm = request.form['Jm']
            motor.Fpv = request.form['Fpv']
            motor.Wn = request.form['Wn']
            motor.Iv = request.form['Iv']
            motor.Vf = request.form['Vf']
            motor.Pav = request.form['Pav']
            motor.Pe = request.form['Pe']
            motor.tipo_partida = request.form['tipo_partida']
            motor.rs = request.form['rs']
            motor.xs = request.form['xs']
            motor.rr = request.form['rr']
            motor.xr = request.form['xr']
            
            # Só atualiza o campo 'fator' se o tipo de partida for 'Partida por Autotransformador'
            if motor.tipo_partida == 'Autotransformador':
                motor.fator = request.form.get('fator')
            
            db.session.commit()
            flash('Motor atualizado com sucesso!', 'success')
            return redirect(url_for('motor'))  # Redireciona de volta para a página de motores

    # Se não for POST, mostra a página de seleção do motor e os campos de edição
    elif request.method == 'GET':
        motor_id = request.args.get('motor_id')  # Usamos GET para carregar o motor
        if motor_id:
            motor = Motor.query.get_or_404(motor_id)
            if motor.user_id != current_user.id:
                flash('Você não tem permissão para editar este motor.', 'danger')
                return redirect(url_for('motor'))

    return render_template('editar_motor.html', motores=motores, motor=motor, user=user, user_profile_image_url=user_profile_image_url, username=username)

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    flash('Você saiu da sua conta.')
    return redirect(url_for('login'))

@app.route('/index')
@login_required
def index():
    user = current_user  # Usar o current_user ao invés de acessar manualmente a sessão
    user_profile_image_url = user.profile_picture
    username = user.username

    # Verifique se o usuário está autenticado corretamente
    if not user.is_authenticated:
        flash('Por favor, faça login para acessar essa página.')
        return redirect(url_for('login'))

    # Pega os dados relacionados ao usuário atual
    motor_data = Motor.query.filter_by(user_id=current_user.id).all()
        
    unique_motor_data = {motor.name: motor for motor in motor_data}.values()

    return render_template(
        'index.html',
        user=user,
        username=username, 
        motor_data=unique_motor_data,
        user_profile_image_url=user_profile_image_url, 
    )



@app.route('/informacoes_complementares')
@login_required
def informacoes_complementares():
    user = current_user  # Usar o current_user ao invés de acessar manualmente a sessão
    user_profile_image_url = user.profile_picture
    username = user.username
    return render_template('informacoes.html',         
        user=user,
        username=username, 
        user_profile_image_url=user_profile_image_url, )



@app.route('/plot', methods=['POST'])
def plot():
    form1_id = request.form.get('form1_id')
    form2_id = request.form.get('form2_id')
    user = current_user
    user_profile_image_url = user.profile_picture
    username = user.username
    graphs = []
    tables = []
    additional_values = {}

    if form1_id == '1' and form2_id == '3':
            # Obtendo dados e gráficos para a exibição
            graph_json1, graph_json2, table_data, additional_values = generate_graph13()
            # Adiciona os gráficos e dados gerados
            graphs.extend([graph_json1, graph_json2])
            # Renderizar template com gráficos interativos
            additional_values.update(additional_values)
            tables.extend([table_data or []])

            return render_template(
                'graph_page_1_3.html',
                user=user,
                graph_json1=graph_json1,
                graph_json2=graph_json2,
                tables=tables,
                additional_values=additional_values,
                username=username,
                user_profile_image_url=user_profile_image_url
            )
            

    # Caso nenhuma combinação seja válida, redireciona para a página inicial com mensagem de erro
    else:
        flash('Combinação inválida de opções selecionadas.')
        return redirect(url_for('index'))

def generate_graph13():
    # Buscar o motor do banco de dados associado ao usuário logado
    motor_id = request.form.get('motor_id')  # Captura o ID do motor
    motor = Motor.query.filter_by(id=motor_id, user_id=current_user.id).first()

    print(f"Motor ID recebido: {motor_id}")
    if not motor:
        flash('Motor não encontrado ou você não tem permissão para acessá-lo.')
        # Retorna valores None e listas vazias para evitar o erro de unpackage
        return None, None, None, None, [], [], {}

       # Obter as informações do banco de dados do motor
    name = motor.name
    rs = motor.rs
    xs = (motor.xs)*1j
    rr = motor.rr
    xr = (motor.xr)*1j
    V = motor.V
    P = motor.P
    Jm = motor.Jm
    Fp = motor.Fp
    rend = motor.rend
    tipo_partida = motor.tipo_partida
    Fpv = motor.Fpv
    Wn = motor.Wn
    Iv = motor.Iv
    Vf = motor.Vf
    Pav = motor.Pav
    Pe = motor.Pe
    f = motor.f
    Jc = float(request.form['Jc'])
    s = np.linspace(1, 0, 1001)

    Tp_Tn = float(request.form.get('Tp_Tn'))
    rr_fator = float(request.form.get('rr_fator'))

    if tipo_partida == 'Direta':
        Vlinha = V
    elif tipo_partida == 'Estrela-triangulo':
        Vlinha = V / np.sqrt(3)
    elif tipo_partida == 'Autotransformador':
        fator = float(request.form.get['fator'])
        Vlinha = V * fator / 100

    Xmag = ((Vlinha/np.sqrt(3))/Iv)*1j
    Zth = ((xs+rs)*Xmag)/(rs+Xmag+xs)

    Tna = (P*1000)/(2*np.pi*(Wn/60))

    # Inicializações
    rr_series = [0]
    teixo_values = []
    found = False

    # Condição para o torque desejado (2.1 * Tn), arredondado
    td = round(Tp_Tn * Tna, 0)
    tolerance = 1e-3  # Tolerância para a condição de parada
    # Contador de iterações
    contador_iteracoes = 0
    contador_iteracoes2 = 0

    # Iteração até encontrar Tn
    while not found:
        contador_iteracoes += 1  # Incrementa o contador
        rr_adjusted = (rr/500) + rr_series[-1]
        rr_series.append(rr_adjusted)

        # Cálculo de Zrotor e corrente
        Zrotor = rr_adjusted + xr + rr

        Vth = ((Vf/np.sqrt(3)) * Xmag) / (rs + Xmag + xs)
        Irotor = Vth / (Zth + Zrotor)

        IROTORA = np.abs(Irotor)

        # Potência do eixo (Peixo)
        Peixo = ((IROTORA** 2) * Zrotor.real) * 3 - Pe

        # Velocidade síncrona (rad/s)
        ws = 2 * np.pi * (Wn/60)

        teixo = round(Peixo / ws, 0)
        teixo_values.append(teixo)

        t2= Zrotor.real

    # Verifica a condição de parada
        if round(abs(teixo - td) ,0) < tolerance:
            contador_iteracoes2 += 1  # Incrementa o contador
            found = True
            Rr_final = rr_adjusted
                        
    Zrotor1=t2-((t2-rr)*((1-s)**(1/rr_fator)))

    Zrotor2 = (Zrotor1/s) +xr
    
    Vth = ((Vlinha/np.sqrt(3)) * Xmag) / (rs + Xmag + xs)

    Irotor2=Vth/(Zth+Zrotor2)

    Ivaziof = (Iv*Fpv)+ complex(0,-Iv* math.sin(math.acos(Fpv)))

    Iestator2 = Ivaziof +Irotor2

    # Calculando as magnitudes de Irotor e Iestator
    IROTORA = abs(Irotor2)
    IESTATORA = abs(Iestator2)

    Pavnovo = Pav *(1-s)
    Pconv = ((IROTORA**2)*(Zrotor2.real))*3
    pw = Pavnovo + Pe
    Peixonova = (((IROTORA** 2) * Zrotor2.real) * 3) - pw
    Pcons = Pconv + ((IESTATORA**2)*(Zth.real*3))
    Pativa = (P)/rend
    Paparente = Pativa/Fp

    ws = (2 * np.pi*(Wn/60))

    Teixo = Peixonova/ws

    In = (P*1000)/(np.sqrt(3)*Fp*rend*Vlinha)

    try:
        # Coletar as entradas fornecidas pelo formulário
        Tnominal = float(request.form['Tnominal'])  # Torque nominal
        Tinicial = float(request.form['Tinicial'])  # Torque inicial
        curve_choice = request.form['curve_choice']  # Escolha da função

        # Avaliar a função de acordo com a escolha do usuário
        if curve_choice == 'linear':
            # Função (Tnominal - Tinicial) * (s-1) + Tinicial
            Tcarga_custom = (Tnominal - Tinicial) * (1-s) + Tinicial
        elif curve_choice == 'quadratic':
            # Função (Tnominal - Tinicial) * (s-1)^2 + Tinicial
            Tcarga_custom = (Tnominal - Tinicial) * (1-s) ** 2 + Tinicial
        elif curve_choice == 'cubic':
            # Função (Tnominal - Tinicial) * (s-1)^3 + Tinicial
            Tcarga_custom = (Tnominal - Tinicial) * (1-s) ** 3 + Tinicial
        else:
            flash('Escolha de função inválida.')
            Tcarga_custom = np.zeros_like(s)  # Em caso de escolha inválida, define uma curva nula

    except Exception as e:
        flash(f"Erro ao calcular a curva de carga: {e}")
        Tcarga_custom = np.zeros_like(s)  # Em caso de erro, define uma curva nula

    Ta = Teixo - Tcarga_custom
    
    # Criar DataFrame df_ta e df_s
    df_ta = pd.DataFrame({'Ta': Ta})
    df_ta['D'] = (df_ta['Ta'] + df_ta['Ta'].shift(-1)) / 2

    df_s = pd.DataFrame({'s': s})
    df_s['N'] = (Jc + Jm) * ws * (df_s['s'] - df_s['s'].shift(-1))
    df_s['Temposubida'] = df_s['N'] / df_ta['D']

    # Ajustando os dois últimos valores da coluna Temposubida para 0
    df_s.loc[df_s.index[-2:], 'Temposubida'] = 0

    # Filtrar os valores de Tempo de Subida para excluir os negativos
    filtered_temposubida1 = df_s[df_s['Temposubida'] > 0]['Temposubida']

    # Verificar se a série filtrada não está vazia
    if not filtered_temposubida1.empty:
        max_value = filtered_temposubida1.max()
        filtered_temposubida = filtered_temposubida1[filtered_temposubida1 != max_value]
    else:
        # Lidar com o caso onde a série filtrada está vazia
        filtered_temposubida = pd.Series([])

    # Calculando a soma dos valores filtrados
    soma_temposubida = filtered_temposubida.sum()

    # Filtrar os valores correspondentes de 's' para manter o alinhamento com filtered_temposubida
    if not filtered_temposubida.empty:
        filtered_s = df_s.loc[filtered_temposubida.index, 's']

        # Garantir que o comprimento de filtered_s e filtered_temposubida sejam iguais
        min_length = min(len(filtered_s), len(filtered_temposubida))
        filtered_s = filtered_s[:min_length]
        filtered_temposubida = filtered_temposubida[:min_length]
    else:
        filtered_s = pd.Series([])


    # Certifique-se de que todos os dados são arrays
    s = np.array(s) if not isinstance(s, np.ndarray) else s
    IROTORA = np.array(IROTORA) if not isinstance(IROTORA, np.ndarray) else IROTORA
    IESTATORA = np.array(IESTATORA) if not isinstance(IESTATORA, np.ndarray) else IESTATORA
    Teixo = np.array(Teixo) if not isinstance(Teixo, np.ndarray) else Teixo
    Pavnovo = np.array(Pavnovo) if not isinstance(Pavnovo, np.ndarray) else Pavnovo
    Pconv = np.array(Pconv) if not isinstance(Pconv, np.ndarray) else Pconv
    Peixonova = np.array(Peixonova) if not isinstance(Peixonova, np.ndarray) else Peixonova
    Pcons = np.array(Pcons) if not isinstance(Pcons, np.ndarray) else Pcons
    ws = np.array(ws) if not isinstance(ws, np.ndarray) else ws
    Tcarga_custom = np.array(Tcarga_custom) if not isinstance(Tcarga_custom, np.ndarray) else Tcarga_custom
    Ta = np.array(Ta) if not isinstance(Ta, np.ndarray) else Ta
    temposubida = np.array(df_s['Temposubida']) if not isinstance(df_s['Temposubida'], np.ndarray) else df_s['Temposubida']

    # Inverta os arrays, se necessário
    s_invertido = np.flip(s)
    T_invertido = np.flip(Teixo)
    Ir_invertido = np.flip(IROTORA)
    Is_invertido = np.flip(IESTATORA)
    Pavnovo_invertido = np.flip(Pavnovo)
    Pconv_invertido = np.flip(Pconv)
    Peixonova_invertido = np.flip(Peixonova)
    Pcons_invertido = np.flip(Pcons)
    Tcarga_custom_invertido = np.flip(Tcarga_custom)
    Ta_invertido = np.flip(Ta)
    Temposubida_invertido = np.flip(temposubida)

    # Validação de comprimentos antes do zip
    lengths = [len(arr) for arr in [
        s_invertido, T_invertido, Ta_invertido, Ir_invertido, Is_invertido,
        Pavnovo_invertido, Pconv_invertido, Peixonova_invertido,
        Pcons_invertido,
        Tcarga_custom_invertido, Temposubida_invertido
    ]]
    if len(set(lengths)) != 1:
        raise ValueError(f"As listas fornecidas têm comprimentos diferentes: {lengths}")

    # Ordenar os dados pelo escorregamento (s_invertido)
    sorted_table_data = sorted(zip(
        s_invertido.tolist(),
        T_invertido.tolist(),
        Ta_invertido.tolist(),
        Ir_invertido.tolist(),
        Is_invertido.tolist(),
        Pavnovo_invertido.tolist(),
        Pconv_invertido.tolist(),
        Peixonova_invertido.tolist(),
        Pcons_invertido.tolist(),
        Tcarga_custom_invertido.tolist(),
        Temposubida_invertido.tolist()
    ), key=lambda x: x[0], reverse=True)

    table_data = []
    for i, (s_val, T_val, Ta_val, Ir_val, Is_val, Pav_val, Pconv_val, Peixo_val, Pcons_val, Tcarga_val, Temposubida_val) in enumerate(sorted_table_data):
        row = {
            'Escorregamento': round(s_val, 4),
            'Corrente Rotórica (A)': round(Ir_val, 4),
            'Corrente Estatórica (A)': round(Is_val, 4),
            'Perdas atrito/ventilação (W)': round(Pav_val, 4),
            'Potência Convertida (W)': round(Pconv_val, 4),
            'Potência no eixo (W)': round(Peixo_val, 4),
            'Potência Consumida (W)': round(Pcons_val, 4),
            'Torque no Eixo (N.m)': round(T_val, 4),
            'Torque na carga (N.m)': round(Tcarga_val, 4),
            'Torque acelerante (N.m)': round(Ta_val, 4),
            'Tempo de aceleração (seg)': safe_value(round(Temposubida_val, 5)) if i < len(Temposubida_invertido) else None,
        }
        # Garantir que apenas as colunas definidas em `columns_order` sejam mantidas
        row = {col: row[col] for col in row}
        table_data.append(row)

    # Preparar os dados para o gráfico
    s_values = [row['Escorregamento'] for row in table_data]
    Tmg_values = [row['Torque no Eixo (N.m)'] for row in table_data]
    Ta_values = [row['Torque acelerante (N.m)'] for row in table_data if row['Torque acelerante (N.m)'] is not None and row['Torque acelerante (N.m)'] > 0]
    # Filtrar valores None em Ta_values e Tmg_values antes de calcular min e max
    Tcarga_values = Tcarga_custom

    # Criar o gráfico
    fig1 = go.Figure()

    # Adicionar as curvas distintas ao gráfico com nomes específicos
    fig1.add_trace(go.Scatter(x=s_values, y=Tmg_values, mode='lines', name='Torque no eixo '))
    fig1.add_trace(go.Scatter(x=s_values, y=Ta_values, mode='lines', name='Torque acelerante', line=dict(color='blue')))
    fig1.add_trace(go.Scatter(x=s_values, y=Tcarga_values, mode='lines', name='Torque da Carga'))

    # Configurar o layout para melhorar a visualização e ativar o grid
    fig1.update_layout(
        title={
        'text': "Torque em função do Escorregamento",
        'x': 0.38,  # Centraliza o título no gráfico
        'xanchor': 'center'
        },    
        xaxis_title="Escorregamento (s)",
        yaxis_title="Torque (N.m)",
        xaxis=dict(
            range=[1, 0],     # Ordenar o eixo x de 1 para 0
            showgrid=True,     # Ativar a grade no eixo x
            gridcolor='lightgrey'  # Cor da grade
        ),
        yaxis=dict(
            showgrid=True,     # Ativar a grade no eixo y
            gridcolor='lightgrey'  # Cor da grade
        ),
        height=510,
        plot_bgcolor='white',     # Define o fundo da área de plotagem como branco
        paper_bgcolor='white'     # Define o fundo geral do gráfico como branco
    )


    # Gráfico 2: Correntes x Escorregamento
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=s_invertido, y=Ir_invertido, mode='lines', name='Corrente Rotórica'))
    fig2.add_trace(go.Scatter(x=s_invertido, y=Is_invertido, mode='lines', name='Corrente Estator '))
    # Configurar o layout para melhorar a visualização e ativar o grid
    fig2.update_layout(
        title={
        'text': "Corrente (A) em função do Escorregamento",
        'x': 0.44,  # Centraliza o título no gráfico
        'xanchor': 'center'
        },    
        xaxis_title="Escorregamento (s)",
        yaxis_title="Corrente (A)",
        xaxis=dict(
            range=[1, 0],     # Ordenar o eixo x de 1 para 0
            showgrid=True,     # Ativar a grade no eixo x
            gridcolor='lightgrey'  # Cor da grade
        ),
        yaxis=dict(
            showgrid=True,     # Ativar a grade no eixo y
            gridcolor='lightgrey'  # Cor da grade
        ),
        height=510,
        plot_bgcolor='white',     # Define o fundo da área de plotagem como branco
        paper_bgcolor='white'     # Define o fundo geral do gráfico como branco
    )

    graph_json1 = json.dumps(fig1, cls=plotly.utils.PlotlyJSONEncoder)
    graph_json2 = json.dumps(fig2, cls=plotly.utils.PlotlyJSONEncoder)


    # Dados adicionais a serem retornados
    additional_values = {
        'Modelo do motor': name,
        'Potência do motor (kW)': P,
        'Modelo da carga': curve_choice,
        'Potência Ativa (kW)': round(Pativa, 4),
        'Potência Aparente (kVA)': round(Paparente, 4),
        'Tensão (V)': V,
        'Tensão de linha (V)': round_values(Vlinha, 4),
        'Rendimento (%)': rend,
        'Resistência do Estator (Ω)': rs,
        'Reatância do Estator (Ω)': convert_complex(xs),
        'Resistência do Rotor (Ω)': rr,
        'Reatância do Rotor (Ω)': convert_complex(xr),
        'Momento de Inércia do Motor (Kg.m²)': Jm,
        'Fator de Potência em Vazio': Fpv,
        'Fator de Potência': Fp,
        'Rotação Nominal do Motor (RPM)': Wn,
        'Corrente em Vazio (A)': Iv,
        'Perdas Atrito/Ventilação (W)': Pav,
        'Perdas ferro magnéticas': Pe,
        'Frequência (Hz)': f,
        'Momento de inércia da Carga (Kg.m²)': Jc,
        'Reatância de Magnetização (Ω)': round(convert_complex(Xmag), 4),
        'Torque Nominal (N.m)': round(Tna, 4),
        'Tipo de partida': tipo_partida,
        'Tempo de Aceleração Total (seg)': round_values(soma_temposubida, 2),
    }


    # Retornar gráficos, tabelas e valores adicionais
    return graph_json1, graph_json2, table_data, additional_values
        
    
@app.route('/export_excel', methods=['POST'])
def export_excel():
    data = request.get_json()

    # Verifique se os dados foram recebidos corretamente
    if not data or 'tables' not in data or not isinstance(data['tables'], list):
        return jsonify({"error": "No table data found or invalid format"}), 400

    additional_values = data.get('additional_values', {})  # Obter valores adicionais, se existirem

    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Itera sobre cada tabela na lista 'tables'
            for index, table in enumerate(data['tables']):
                if 'data' not in table or 'sheet_name' not in table:
                    return jsonify({"error": f"Missing 'data' or 'sheet_name' in table at index {index}"}), 400

                # Converte os dados da tabela em DataFrame
                try:
                    df = pd.DataFrame(table['data'])
                except Exception as e:
                    return jsonify({"error": f"Failed to convert table {index} to DataFrame: {str(e)}"}), 500

                # Escreve o DataFrame na planilha correspondente
                df.to_excel(writer, index=False, sheet_name=table['sheet_name'])

            # Adicionar a aba com os valores adicionais
            if additional_values:
                df_additional = pd.DataFrame(additional_values.items(), columns=['Parâmetro', 'Valor'])
                df_additional.to_excel(writer, index=False, sheet_name='Valores Adicionais')

            # Estilos para bordas finas e alinhamento centralizado
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            alignment_center = Alignment(horizontal="center", vertical="center")

            # Ajustar colunas e aplicar estilos às células
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                # Ajusta a largura das colunas com base no conteúdo
                for col_cells in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in col_cells if cell.value)
                    col_letter = col_cells[0].column_letter
                    worksheet.column_dimensions[col_letter].width = max_length + 2

                # Aplica estilo ao cabeçalho
                for cell in worksheet[1]:  # Primeira linha (cabeçalho)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = alignment_center

                # Aplica bordas e alinhamento a todas as células
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = alignment_center  # Centralizar o conteúdo
                        cell.border = thin_border         # Adicionar bordas às células

        # Reseta o ponteiro para o início do arquivo
        output.seek(0)

    except Exception as e:
        return jsonify({"error": f"Failed to create Excel file: {str(e)}"}), 500

    # Retorna o arquivo Excel como um download
    return send_file(output, download_name='dados_exportados.xlsx', as_attachment=True)

if __name__ == '__main__':
    create_tables()  # Cria as tabelas no banco de dados
    create_master_user()  # Cria o usuário administrador
    run_server()
    app.run(debug=True)